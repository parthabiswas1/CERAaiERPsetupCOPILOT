from fastapi import FastAPI, Depends, HTTPException, Request, File, UploadFile
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from uuid import uuid4
import sqlite3, os, hashlib, time, json, pathlib
from typing import Dict
from ceraai.tools import RAGTool, RulesTool, ERPConnector, AuditTool
from ceraai.agents import InterviewAgent, ValidatorAgent, MapperAgent, ExecutorAgent, AuditorAgent
import random
from fastapi.middleware.cors import CORSMiddleware
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime


DB_PATH = os.getenv("DB_PATH", str(pathlib.Path("/tmp/ceraai.db")))

FRONTEND_ORIGIN = os.getenv("FRONTEND_ORIGIN", "*")  # set to your UI origin later

UPLOAD_ROOT = pathlib.Path("/tmp/ceraai_uploads")
UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="CERAai ERP Setup Copilot - MVP")
security = HTTPBasic()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[FRONTEND_ORIGIN],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def boot():
    init_db()  # make sure rules/runs/idem tables exist at process start


@app.middleware("http")
async def ensure_run_id(request: Request, call_next):
    rid = request.headers.get("x-run-id") or str(uuid4())
    request.state.run_id = rid
    resp = await call_next(request)
    resp.headers["X-Run-ID"] = rid
    return resp


rag = RAGTool()
rules_tool = RulesTool(DB_PATH)
interview_agent = InterviewAgent()
validator_agent = ValidatorAgent()
executor_agent  = ExecutorAgent()
erp_connector   = ERPConnector()
auditor_agent   = AuditorAgent()
audit_tool      = AuditTool()
mapper_agent    = MapperAgent()


def _run_dir(run_id: str) -> pathlib.Path:
    d = UPLOAD_ROOT / run_id
    d.mkdir(parents=True, exist_ok=True)
    return d


# --- state persistence---


def init_db():
    con = sqlite3.connect(DB_PATH); cur = con.cursor()
    # base table (old installs won’t have new columns)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS runs (
      run_id TEXT PRIMARY KEY,
      inputs TEXT DEFAULT '{}',
      missing TEXT DEFAULT '[]',
      mapped TEXT,
      result TEXT
    )""")
    # migrate: add new columns if missing
    def _add(col, ddl):
        try: cur.execute(f"ALTER TABLE runs ADD COLUMN {col} {ddl}")
        except sqlite3.OperationalError: pass
    _add("gating", "TEXT DEFAULT '{}'")
    _add("phase", "TEXT DEFAULT 'gating'")
    _add("required_fields", "TEXT")
    _add("current_ask_for", "TEXT")
    _add("created_at", "TEXT DEFAULT (strftime('%Y-%m-%dT%H:%M:%fZ','now'))")
    _add("updated_at", "TEXT")
    con.commit(); con.close()

def _jdump(x):
    return json.dumps(x) if x is not None else None

def _jload(s, default):
    try:
        return json.loads(s) if s not in (None, "") else default
    except Exception:
        return default

def load_state(run_id: str) -> dict:
    con = sqlite3.connect(DB_PATH); cur = con.cursor()
    cur.execute("""SELECT inputs, missing, mapped, result, gating, phase, required_fields, current_ask_for
                   FROM runs WHERE run_id=?""", (run_id,))
    row = cur.fetchone(); con.close()
    if not row:
        return {
            "run_id": run_id,
            "inputs": {}, "missing": [], "mapped": None, "result": None,
            "gating": {}, "phase": "gating", "required_fields": None, "current_ask_for": None,
        }
    inputs, missing, mapped, result, gating, phase, required_fields, current_ask_for = row
    return {
        "run_id": run_id,
        "inputs": _jload(inputs, {}), "missing": _jload(missing, []),
        "mapped": _jload(mapped, None), "result": _jload(result, None),
        "gating": _jload(gating, {}), "phase": phase or "gating",
        "required_fields": _jload(required_fields, None), "current_ask_for": current_ask_for,
    }

def save_state(run_id: str, **st):
    con = sqlite3.connect(DB_PATH); cur = con.cursor()
    # ensure row exists
    cur.execute("INSERT OR IGNORE INTO runs (run_id) VALUES (?)", (run_id,))

    # discover existing columns, build UPDATE accordingly
    cols_info = cur.execute("PRAGMA table_info(runs)").fetchall()
    existing = {c[1] for c in cols_info}

    values_map = {
        "inputs": _jdump(st.get("inputs", {})),
        "missing": _jdump(st.get("missing", [])),
        "mapped": _jdump(st.get("mapped", None)),
        "result": _jdump(st.get("result", None)),
        "gating": _jdump(st.get("gating", {})),
        "phase": st.get("phase", "gating"),
        "required_fields": _jdump(st.get("required_fields", None)),
        "current_ask_for": st.get("current_ask_for"),
        "updated_at": datetime.utcnow().isoformat() + "Z",
    }

    set_cols, set_vals = [], []
    for col, val in values_map.items():
        if col in existing:
            set_cols.append(f"{col}=?")
            set_vals.append(val)
    set_vals.append(run_id)

    cur.execute(f"UPDATE runs SET {', '.join(set_cols)} WHERE run_id=?", set_vals)
    con.commit(); con.close()



def _conn(): return sqlite3.connect(DB_PATH)

def load_state(run_id: str) -> dict:
    con = sqlite3.connect(DB_PATH); cur = con.cursor()
    try:
        cur.execute("SELECT inputs,missing,mapped,result FROM runs WHERE run_id=?", (run_id,))
        row = cur.fetchone()
    except sqlite3.OperationalError as e:
        # table missing -> initialize once and return empty state
        if "no such table" in str(e):
            con.close()
            init_db()
            return {"inputs": {}, "missing": [], "mapped": None, "result": None}
        con.close()
        raise
    con.close()
    if not row:
        return {"inputs": {}, "missing": [], "mapped": None, "result": None}
    inputs, missing, mapped, result = row
    return {
        "inputs": json.loads(inputs or "{}"),
        "missing": json.loads(missing or "[]"),
        "mapped": json.loads(mapped or "null"),
        "result": json.loads(result or "null"),
    }


def save_state(run_id: str, **patch):
    st = load_state(run_id)
    st.update(patch)
    con=_conn(); cur=con.cursor()
    cur.execute("""INSERT INTO runs(run_id,inputs,missing,mapped,result,updated_at)
                   VALUES(?,?,?,?,?,?)
                   ON CONFLICT(run_id) DO UPDATE SET
                     inputs=excluded.inputs, missing=excluded.missing,
                     mapped=excluded.mapped, result=excluded.result,
                     updated_at=excluded.updated_at""",
        (run_id,
         json.dumps(st["inputs"]), json.dumps(st["missing"]),
         json.dumps(st["mapped"]), json.dumps(st["result"]), int(time.time())))
    con.commit(); con.close()

def idem_get(action: str, key: str):
    con=_conn(); cur=con.cursor()
    cur.execute("SELECT resp FROM idem WHERE action=? AND key=?", (action,key))
    row=cur.fetchone(); con.close()
    return json.loads(row[0]) if row else None

def idem_put(action: str, key: str, resp: dict):
    con=_conn(); cur=con.cursor()
    cur.execute("INSERT OR REPLACE INTO idem(action,key,resp) VALUES(?,?,?)",
                (action,key,json.dumps(resp)))
    con.commit(); con.close()



def basic_ok(creds: HTTPBasicCredentials):
    return creds.username == os.getenv("DEMO_USER", "demo") and creds.password == os.getenv("DEMO_PASS", "demo")

def satisfied(inputs: dict) -> bool:
    ctry = (inputs.get("country") or "").upper()
    if ctry == "US":
        base = all(inputs.get(k) for k in
                   ["legal_name","country","address_line1","city","state_region","postal_code","ein"])
        if not base: return False
        if str(inputs.get("employees_in_CA")).lower() == "true":
            return bool(inputs.get("ca_edd_id"))
        return True
    # Non-US: minimal example
    return all(inputs.get(k) for k in
               ["legal_name","country","address_line1","city","state_region","postal_code","tax_id"])


GATING_ORDER = ["country","employ_in_country","sells_in_country","has_tax_id","regulatory_category"]

def get_country_pack(ctry: str | None):
    # TODO: replace with Pinecone RAG fetch; fallback pack for now
    return {
      "country": (ctry or "US").upper(),
      "base_fields": ["legal_name","country","address_line1","city","state_region","postal_code","ein"],
      "gating": [
        {"key":"employ_in_country","type":"bool","on_true_add":["employer_registration_id"]},
        {"key":"sells_in_country","type":"bool","on_true_add":["indirect_tax_id"]},
        {"key":"has_tax_id","type":"bool"},
        {"key":"regulatory_category","type":"enum","options":["none","npo","gov","fsi"],"map":{"fsi":["regulator_code"]}}
      ],
      "field_hints": {"ein":"Format NN-NNNNNNN.","indirect_tax_id":"Sales/VAT permit if selling."}
    }

def normalize_country(text: str) -> str | None:
    t = text.strip().lower()
    if t in {"us","usa","united states","united states of america"}:
        return "US"
    return t.upper() if len(t) == 2 else None

def parse_bool(text: str) -> bool | None:
    t = text.strip().lower()
    if t in {"true","yes","y"}: return True
    if t in {"false","no","n"}: return False
    return None


def _reg_opts(pack: dict) -> set:
    for g in pack.get("gating", []):
        if g.get("key") == "regulatory_category":
            return set(g.get("options", []))
    return {"none","npo","gov","fsi"}  # fallback

def validate_gating_answer(key: str, text: str, pack: dict):
    if key == "country":
        c = normalize_country(text)
        return (True, {"country": c}) if c else (False, "Reply with a valid 2-letter ISO code (e.g., US).")
    if key in {"employ_in_country","sells_in_country","has_tax_id"}:
        b = parse_bool(text)
        return (True, {key: b}) if b is not None else (False, "Reply exactly: true or false.")
    if key == "regulatory_category":
        opts = _reg_opts(pack)
        v = text.strip().lower()
        return (True, {key: v}) if v in opts else (False, f"Choose one of: {', '.join(sorted(opts))}.")
    return (False, "Unsupported gating key.")

def next_gating_key(st: dict, pack: dict) -> str | None:
    g = st.get("gating", {})
    for k in GATING_ORDER:
        if g.get(k) in (None, ""):
            return k
    return None

def gating_question(key: str, pack: dict) -> str:
    # (RAG phrasing later) deterministic, value-seeking fallback
    fallback = {
      "country": "Which country is this legal entity in? Reply with the 2-letter ISO code (e.g., US).",
      "employ_in_country": "Will this entity employ staff in this country at go-live? Reply true or false.",
      "sells_in_country": "Will this entity sell goods/services in this country? Reply true or false.",
      "has_tax_id": "Do you already have a national tax ID for this entity? Reply true or false.",
      "regulatory_category": "Regulatory category? Reply one of: none, npo, gov, fsi."
    }
    return fallback.get(key, f"Provide value for {key}.")

def derive_required_fields(pack: dict, gating: dict) -> list[str]:
    req = list(pack["base_fields"])
    for g in pack["gating"]:
        k = g["key"]; v = gating.get(k)
        if v is True and "on_true_add" in g:
            req += g["on_true_add"]
        if isinstance(v, str) and g.get("type") == "enum":
            req += (g.get("map") or {}).get(v, [])
    # de-dup while preserving order
    return list(dict.fromkeys(req))


def parse_address_block(txt: str) -> dict | None:
    parts = [p.strip() for p in txt.split(",")]
    if len(parts) < 4:
        return None
    return {
        "address_line1": parts[0],
        "city": parts[1],
        "state_region": parts[2],
        "postal_code": parts[3],
    }


def satisfied(inputs: dict) -> bool:
    ctry = (inputs.get("country") or "").upper()
    if ctry == "US":
        base = all(inputs.get(k) for k in [
            "legal_name","country","address_line1","city","state_region","postal_code","ein"
        ])
        return bool(base)
    # Non-US minimal (adjust per pack when RAG is wired)
    return all(inputs.get(k) for k in [
        "legal_name","country","address_line1","city","state_region","postal_code","tax_id"
    ])




@app.get("/")
def root():
    return {"service": "CERAaiERPsetupCOPILOT", "status": "ok"}

@app.get("/health")
def health():
    return {"ok": True}

@app.get("/secure-check")
def secure_check(credentials: HTTPBasicCredentials = Depends(security)):
    if basic_ok(credentials): return {"access": "granted"}
    raise HTTPException(status_code=401, detail="Unauthorized")

@app.post("/validate")
def validate(payload: dict, request: Request, credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401, detail="Unauthorized")
    res = validator_agent.validate(payload, rules_tool)
    return {"run_id": request.state.run_id, **res}

@app.get("/state")
def get_state(request: Request, credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401)
    return {"run_id": request.state.run_id, **load_state(request.state.run_id)}

@app.post("/execute")
def execute(payload: dict | None = None, request: Request = None,
           credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401)
    st = load_state(request.state.run_id)

    # prefer mapped from state if payload not supplied
    work = payload or st.get("mapped")
    if not work:
        raise HTTPException(status_code=400, detail="No mapped payload. Call /map first or pass payload.")

    # idempotency
    idem_key = request.headers.get("idempotency-key")
    if idem_key:
        cached = idem_get("execute", idem_key)
        if cached: return {"run_id": request.state.run_id, **cached}

    res = executor_agent.execute(work, erp_connector)
    st["result"] = res; save_state(request.state.run_id, **st)

    if idem_key: idem_put("execute", idem_key, res)
    return {"run_id": request.state.run_id, **res}

LOG_FILE = "audit_log.jsonl"

#def write_audit(entry: dict):
#    entry["timestamp"] = int(time.time())
#    with open(LOG_FILE, "a") as f:
#        f.write(json.dumps(entry) + "\n")

@app.post("/audit")
def audit(event: dict, request: Request, credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401, detail="Unauthorized")
    event = {**event, "run_id": request.state.run_id}
    return auditor_agent.record(event, audit_tool)

@app.get("/audit/logs")
def audit_logs(run_id: str | None = None, limit: int = 200,
               credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401, detail="Unauthorized")
    data = auditor_agent.fetch(audit_tool, limit)["logs"]
    return {"logs": [e for e in data if not run_id or e.get("run_id") == run_id]}


@app.get("/interview/next")
def interview_next(request: Request, credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401)
    st = load_state(request.state.run_id)
    st.setdefault("phase","gating")
    st.setdefault("gating",{})
    st.setdefault("inputs",{})
    pack = get_country_pack(st["gating"].get("country") or st["inputs"].get("country"))

    # GATING PHASE
    if st["phase"]=="gating":
        key = next_gating_key(st, pack)
        if key:
            # RAG phrasing later; for now deterministic, value-seeking
            fallback = {
              "country": "Which country is this legal entity in? Reply with the 2-letter ISO code (e.g., US).",
              "employ_in_country": "Will this entity employ staff in this country at go-live? Reply true or false.",
              "sells_in_country": "Will this entity sell goods/services in this country? Reply true or false.",
              "has_tax_id": "Do you already have a national tax ID for this entity? Reply true or false.",
              "regulatory_category": "Regulatory category? Reply one of: none, npo, gov, fsi."
            }
            q = fallback.get(key, f"Provide value for {key}.")
            st["current_ask_for"] = key
            save_state(request.state.run_id, **st)
            return {"run_id": request.state.run_id, "phase":"gating", "ask_for":key, "complete":False, "question":q}

        # all gating answered → derive required fields and switch to values
        req = derive_required_fields(pack, st["gating"])
        st["required_fields"] = req
        res = validator_agent.validate(st["inputs"], rules_tool)
        st["missing"] = res["missing"]
        st["phase"]="values"
        save_state(request.state.run_id, **st)

    # VALUES PHASE
    if satisfied(st["inputs"]):
        return {"run_id": request.state.run_id, "phase":"values", "complete":True,
                "question":"I have enough to generate your Legal Entity template."}

    missing = [m["field"] for m in st.get("missing",[])]
    if any(x in missing for x in ["address_line1","city","state_region","postal_code"]):
        ask_for = "address_block"
        q = "Provide the legal address as: line1, city, state/region, postal code. Example: 123 Main St, San Jose, CA, 95110."
    else:
        ask_for = missing[0] if missing else "confirmation"
        hint = (pack.get("field_hints",{}) or {}).get(ask_for,"")
        q = f"Provide {ask_for}. {hint}".strip()

    st["current_ask_for"] = ask_for
    save_state(request.state.run_id, **st)
    return {"run_id": request.state.run_id, "phase":"values", "ask_for":ask_for, "complete":False, "question":q}

@app.post("/interview/answer")
def interview_answer(payload: dict, request: Request, credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401)
    text = (payload or {}).get("text","").strip()
    if not text: raise HTTPException(status_code=400, detail="Provide 'text'.")

    st = load_state(request.state.run_id)
    st.setdefault("phase","gating"); st.setdefault("gating",{}); st.setdefault("inputs",{})
    pack = get_country_pack(st["gating"].get("country") or st["inputs"].get("country"))

    # GATING
    if st["phase"]=="gating":
        key = next_gating_key(st, pack)
        if not key:
            st["phase"]="values"  # safety
        else:
            ok, result = validate_gating_answer(key, text, pack)
            if not ok:
                return {"run_id": request.state.run_id, "phase":"gating", "accepted":False, "message":result}
            st["gating"].update(result)
            save_state(request.state.run_id, **st)
            # Auto-ask next
            nxt = interview_next(request, credentials)
            return {"run_id": request.state.run_id, "phase":"gating", "accepted":True, "next":nxt}

    # VALUES
    ask_for = st.get("current_ask_for")
    extracted = {}

    if ask_for == "address_block":
        addr = parse_address_block(text)
        if not addr:
            return {"run_id": request.state.run_id, "phase":"values", "accepted":False,
                    "message":"Please provide address as: line1, city, state/region, postal code."}
        extracted.update(addr)

    elif ask_for == "ein":
        import re
        m = re.search(r"\b\d{2}-\d{7}\b", text)
        if not m: return {"run_id": request.state.run_id, "phase":"values", "accepted":False,
                          "message":"EIN format must be NN-NNNNNNN (e.g., 12-3456789)."}
        extracted["ein"] = m.group(0)

    elif ask_for == "ca_edd_id":
        import re
        m = re.search(r"\b\d{3}-\d{4}\b", text)
        if not m: return {"run_id": request.state.run_id, "phase":"values", "accepted":False,
                          "message":"Provide CA EDD Employer Account like 123-4567."}
        extracted["ca_edd_id"] = m.group(0)

    elif ask_for == "legal_name":
        extracted["legal_name"] = text

    elif ask_for == "tax_id":
        extracted["tax_id"] = text

    else:
        # Generic: set the field to the raw text
        if ask_for and ask_for not in {"confirmation"}:
            extracted[ask_for] = text

    # Merge + validate
    st["inputs"].update(extracted)
    res = validator_agent.validate(st["inputs"], rules_tool)
    st["missing"] = res["missing"]
    is_ready = satisfied(st["inputs"])
    save_state(request.state.run_id, **st)

    if is_ready:
        return {"run_id": request.state.run_id, "phase":"values", "accepted":True, "complete":True,
                "message":"All required data captured. You can download the template now."}

    # Auto-ask next
    nxt = interview_next(request, credentials)
    return {"run_id": request.state.run_id, "phase":"values", "accepted":True, "complete":False,
            "next":nxt, "missing":st["missing"], "extracted":extracted}


@app.post("/map")
def map_to_fusion(payload: dict | None = None, request: Request = None,
                  credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401)
    st = load_state(request.state.run_id)
    if payload:  # allow overriding/adding
        st["inputs"].update(payload)

    res = validator_agent.validate(st["inputs"], rules_tool)
    if res["status"] != "ok":
        save_state(request.state.run_id, inputs=st["inputs"], missing=res["missing"])
        raise HTTPException(status_code=422, detail={"missing": res["missing"]})

    mapped = mapper_agent.map_to_fusion(st["inputs"])
    st["mapped"] = mapped; save_state(request.state.run_id, **st)
    return {"run_id": request.state.run_id, "status":"ok", "mapped": mapped}

# --- Dynamic XLSX template builder ---
def build_template_xlsx(country: str = "US", required_fields: list[str] | None = None) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "LegalEntity"

    ctry = (country or "US").upper()

    # Columns: use derived required_fields if present; otherwise fall back
    if required_fields and isinstance(required_fields, list) and required_fields:
        cols = required_fields
    else:
        base_cols = ["legal_name","country","address_line1","city","state_region","postal_code"]
        cols = base_cols + (["ein"] if ctry == "US" else ["tax_id"])

    ws.append(cols)

    # Guidance/demo row mapped by header name (only fills what exists)
    demo_map = {
        "legal_name": "ABC, Inc.",
        "country": ctry,
        "address_line1": "123 Main St",
        "city": "San Jose",
        "state_region": "CA",
        "postal_code": "95110",
        "ein": "12-3456789",
        "tax_id": "TAX-XXX",
        "employer_registration_id": "",
        "indirect_tax_id": "",
        "regulator_code": "",
    }
    ws.append([demo_map.get(c, "") for c in cols])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.get("/template/draft")
def template_draft(request: Request, country: str | None = None,
                   credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials):
        raise HTTPException(status_code=401, detail="Unauthorized")

    st = load_state(request.state.run_id)
    ctry = (country
            or st.get("gating", {}).get("country")
            or st.get("inputs", {}).get("country")
            or "US").upper()
    fields = st.get("required_fields")  # set after gating is complete

    xlsx = build_template_xlsx(ctry, fields)
    return StreamingResponse(
        xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="legal_entity_template_{ctry}.xlsx"'}
    )



@app.post("/files/upload")
async def files_upload(request: Request, file: UploadFile = File(...),
                       credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401, detail="Unauthorized")
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx accepted")
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 10MB)")
    p = _run_dir(request.state.run_id) / "uploaded.xlsx"
    p.write_bytes(contents)
    return {"run_id": request.state.run_id, "stored": True, "path": str(p)}


@app.post("/files/validate")
def files_validate(request: Request, credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401, detail="Unauthorized")
    p = _run_dir(request.state.run_id) / "uploaded.xlsx"
    if not p.exists():
        raise HTTPException(status_code=400, detail="No uploaded file for this run. POST /files/upload first.")

    wb = load_workbook(p)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    issues = []

    # required columns (US as default); country from state if present
    st = load_state(request.state.run_id)
    ctry = (st["inputs"].get("country") or "US").upper()
    base_req = ["legal_name","country","address_line1","city","state_region","postal_code"]
    req = base_req + (["ein"] if ctry=="US" else ["tax_id"])

    for col in req:
        if col not in headers:
            issues.append({"row": 1, "field": col, "error": "Missing required column"})

    # per-row empties for first data row (row 2)
    hdr_idx = {h:i for i,h in enumerate(headers)}
    if not issues and ws.max_row >= 2:
        r = 2
        for col in req:
            cell = ws.cell(row=r, column=hdr_idx[col]+1)
            if cell.value in (None, ""):
                issues.append({"row": r, "field": col, "error": "Value required"})

    # produce review.xlsx with Errors column + highlight
    if "Errors" not in headers:
        ws.cell(row=1, column=len(headers)+1, value="Errors")
        headers.append("Errors")
    red = PatternFill(start_color="FFFECACA", end_color="FFFECACA", fill_type="solid")
    for item in issues:
        r = item["row"]
        # append message into Errors column
        err_col = len(headers)
        curr = ws.cell(row=r, column=err_col).value or ""
        msg = f"{item['field']}: {item['error']}"
        ws.cell(row=r, column=err_col, value=(curr + ("; " if curr else "") + msg))
        # highlight offending cell if we can
        if item["field"] in hdr_idx:
            ws.cell(row=r, column=hdr_idx[item["field"]]+1).fill = red

    review_path = _run_dir(request.state.run_id) / "review.xlsx"
    wb.save(review_path)

    return {"run_id": request.state.run_id, "issues": issues,
            "review_download": "/files/review"}


@app.get("/files/review")
def files_review(request: Request, credentials: HTTPBasicCredentials = Depends(security)):
    if not basic_ok(credentials): raise HTTPException(status_code=401, detail="Unauthorized")
    review_path = _run_dir(request.state.run_id) / "review.xlsx"
    if not review_path.exists():
        raise HTTPException(status_code=404, detail="No review file for this run.")
    return FileResponse(review_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        filename="review.xlsx")



